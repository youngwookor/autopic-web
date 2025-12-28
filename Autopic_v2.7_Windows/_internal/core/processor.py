# -*- coding: utf-8 -*-
"""
메인 프로세서
=============
- 전체 파이프라인 통합
- 기본 + 화보 이미지 생성 지원
"""

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Callable
from dataclasses import dataclass, field

from .settings import get_manager, get_settings
from .gemini_client import GeminiClient
from .claude_client import ClaudeClient
from .image_generator import ImageGenerator, get_category_group
from .product_analyzer import ProductAnalyzer, ProductInfo, ImageSorter
from .seo_generator import SEOGenerator, SEOContent


@dataclass
class ProcessResult:
    product_code: str
    folder_path: str
    brand: str = ""
    category1: str = ""
    category2: str = ""
    product_name: str = ""
    gender: str = ""
    detected_gender: str = ""  # AI 감지된 성별
    user_gender: str = ""      # 사용자 선택 성별
    seo_title: str = ""
    seo_description: str = ""
    seo_keywords: str = ""
    success: bool = False
    error_message: str = ""
    images_generated: List[str] = field(default_factory=list)
    brand_is_new: bool = False
    needs_review: bool = False
    
    def get_excel_gender(self) -> str:
        """엑셀용 성별 문자열 생성"""
        detected = self.detected_gender or "공용"
        user = self.user_gender or "auto"
        
        if user == "auto":
            # 자동 감지 모드
            if detected == "공용":
                return "남여공용"
            else:
                return f"{detected},남여공용"
        else:
            # 사용자 지정 모드
            if user == detected:
                return f"{user},남여공용"
            else:
                return f"{user},남여공용"
    
    def to_excel_row(self) -> dict:
        return {
            "상품코드": self.product_code, "브랜드": self.brand,
            "1차카테고리": self.category1, "2차카테고리": self.category2,
            "상품명": self.product_name, "성별": self.get_excel_gender(),
            "SEO제목": self.seo_title, "SEO설명": self.seo_description, "SEO키워드": self.seo_keywords,
        }


class MainProcessor:
    def __init__(self, callback: Callable[[str], None] = None):
        self.callback = callback or print
        self.manager = get_manager()
        self.settings = self.manager.settings
        self.gemini_client: Optional[GeminiClient] = None
        self.claude_client: Optional[ClaudeClient] = None
        self.image_generator: Optional[ImageGenerator] = None
        self.product_analyzer: Optional[ProductAnalyzer] = None
        self.seo_generator: Optional[SEOGenerator] = None
        self.image_sorter: Optional[ImageSorter] = None
        self.is_running = False
        self.should_stop = False
        self.auto_add_new_brands = False
        self.existing_excel_data: Dict[str, dict] = {}  # 기존 엑셀 데이터 캐시
    
    def load_existing_excel_data(self, work_folder: Path):
        """엑셀에서 기존 분석 데이터 로드"""
        self.existing_excel_data = {}
        excel_path = work_folder / f"{work_folder.name}_result.xlsx"
        
        if not excel_path.exists():
            return
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    row_data[header] = ws.cell(row=row_idx, column=col_idx).value or ""
                
                product_code = row_data.get("상품코드", "")
                if product_code:
                    self.existing_excel_data[str(product_code)] = row_data
            
            self.callback(f"📊 기존 엑셀 데이터 로드: {len(self.existing_excel_data)}개")
        except Exception as e:
            self.callback(f"⚠️ 엑셀 로드 실패: {e}")
    
    def get_existing_data(self, product_code: str) -> Optional[dict]:
        """기존 분석 데이터 확인 - 상품명+카테고리 있으면 반환"""
        data = self.existing_excel_data.get(str(product_code))
        if not data:
            return None
        
        # 필수 필드 확인 (상품명 + 1차카테고리)
        product_name = data.get("상품명", "")
        category1 = data.get("1차카테고리", "")
        
        if product_name and category1:
            return data
        return None
    
    def initialize(self) -> bool:
        try:
            if self.settings.GEMINI_API_KEYS:
                self.gemini_client = GeminiClient(
                    api_keys=self.settings.GEMINI_API_KEYS,
                    retry_count=self.settings.GEMINI_RETRY_COUNT,
                    retry_delay=self.settings.GEMINI_RETRY_DELAY
                )
                self.image_generator = ImageGenerator(self.gemini_client, upscale_factor=self.settings.UPSCALE_FACTOR)
                self.callback("✅ Gemini API 초기화 완료")
            else:
                self.callback("⚠️ Gemini API 키가 없습니다")
                return False
            
            if self.settings.CLAUDE_API_KEY:
                self.claude_client = ClaudeClient(self.settings.CLAUDE_API_KEY)
                self.product_analyzer = ProductAnalyzer(self.claude_client)
                self.seo_generator = SEOGenerator(self.claude_client)
                self.image_sorter = ImageSorter(self.claude_client)
                self.callback("✅ Claude API 초기화 완료")
            else:
                self.callback("⚠️ Claude API 키가 없습니다 - 상품 분석 비활성화")
            
            return True
        except Exception as e:
            self.callback(f"❌ 초기화 실패: {e}")
            return False
    
    def process_folder(self, folder_path: Path, product_code: str, options: dict = None) -> ProcessResult:
        options = options or {}
        result = ProcessResult(product_code=product_code, folder_path=str(folder_path))
        result.user_gender = options.get("gender", "auto")
        
        try:
            self.callback(f"\n{'='*50}")
            self.callback(f"📦 {product_code} 처리 시작")
            self.callback(f"{'='*50}")
            
            # 0. 기존 엑셀 데이터 확인 (AI 분석 스킵 여부)
            existing_data = self.get_existing_data(product_code)
            use_existing_data = existing_data is not None
            
            if use_existing_data:
                self.callback(f"  📋 기존 분석 데이터 사용 (AI 분석 스킵)")
                result.brand = existing_data.get("브랜드", "")
                result.category1 = existing_data.get("1차카테고리", "")
                result.category2 = existing_data.get("2차카테고리", "")
                result.product_name = existing_data.get("상품명", "")
                result.seo_title = existing_data.get("SEO제목", "")
                result.seo_description = existing_data.get("SEO설명", "")
                result.seo_keywords = existing_data.get("SEO키워드", "")
                # 성별 파싱 ("여성,남여공용" -> "여성")
                gender_str = existing_data.get("성별", "여성")
                result.detected_gender = gender_str.split(",")[0] if gender_str else "여성"
            
            # 1. 텍스트 파일 찾기 (기존 데이터 없을 때만)
            text_content = ""
            if not use_existing_data:
                text_content = self._find_and_read_text(folder_path)
                if not text_content:
                    self.callback("  ⚠️ 텍스트 파일 없음")
            
            # 2. 이미지 정렬 (기존 데이터 없을 때만)
            if not use_existing_data and self.image_sorter:
                self.image_sorter.sort_images(folder_path, self.callback)
            
            # 3. 정면 이미지 찾기
            front_image = self._find_front_image(folder_path)
            if not front_image:
                result.error_message = "정면 이미지(5.jpg)를 찾을 수 없습니다"
                self.callback(f"  ❌ {result.error_message}")
                return result
            
            self.callback(f"  📷 정면 이미지: {front_image.name}")
            
            # 4. 후면 이미지 찾기
            back_image = self._find_back_image(folder_path)
            if back_image:
                self.callback(f"  📷 후면 이미지: {back_image.name}")
            
            # 5. 카테고리/성별 감지 (기존 데이터 없을 때만)
            if not use_existing_data:
                category1, category2, detected_gender = "", "", "여성"
                if self.image_generator:
                    category1, category2, detected_gender = self.image_generator.detect_category(front_image, self.callback)
                result.detected_gender = detected_gender
            
            # 6. 상품 분석 (기존 데이터 없을 때만 - 업종에 따라 브랜드 분석 ON/OFF)
            if not use_existing_data:
                use_brand = self.settings.USE_BRAND
                
                if self.product_analyzer:
                    product_info = self.product_analyzer.analyze(front_image, text_content, self.callback, analyze_brand=use_brand)
                    if use_brand:
                        if product_info.brand_is_new:
                            result.brand_is_new = True
                            if self.auto_add_new_brands:
                                self.manager.brands.add(product_info.brand)
                                self.manager.save_brands()
                                self.callback(f"  ✅ 새 브랜드 자동 추가: {product_info.brand}")
                        result.brand = product_info.brand
                    else:
                        result.brand = ""  # 일반 패션은 브랜드 없음
                    
                    result.category1 = product_info.category1 or category1
                    result.category2 = product_info.category2 or category2
                    result.product_name = product_info.product_name
                    
                    if product_info.gender:
                        result.detected_gender = product_info.gender
                    
                    if product_info.category_uncertain:
                        result.needs_review = True
                else:
                    result.category1 = category1
                    result.category2 = category2
            
            # 7. 성별 결정 (이미지 생성용)
            user_gender = options.get("gender", "auto")
            if user_gender == "auto":
                gender_for_image = result.detected_gender or "여성"
            else:
                gender_for_image = user_gender
            
            result.gender = gender_for_image
            self.callback(f"  👤 성별: {gender_for_image} (감지: {result.detected_gender}, 설정: {user_gender})")
            
            # 8. 이미지 생성 옵션 (기본 + 화보)
            if self.image_generator:
                output_dir = folder_path / "output"
                
                gen_results = self.image_generator.generate_all(
                    input_image_path=front_image,
                    output_dir=output_dir,
                    gender=gender_for_image,
                    category1=result.category1,
                    category2=result.category2,
                    basic_product=options.get("basic_product", True),
                    basic_model=options.get("basic_model", True),
                    editorial_product=options.get("editorial_product", False),
                    editorial_model=options.get("editorial_model", False),
                    callback=self.callback,
                )
                
                for image_type, files in gen_results.items():
                    for f in files:
                        result.images_generated.append(str(f))
            
            # 9. SEO 생성 (기존 데이터 없거나 SEO가 비어있을 때만)
            if not use_existing_data or not result.seo_title:
                if self.seo_generator and result.brand and result.category1:
                    self.callback("  📝 SEO 콘텐츠 생성 중...")
                    seo = self.seo_generator.generate(
                        ProductInfo(brand=result.brand, category1=result.category1, category2=result.category2,
                                   product_name=result.product_name, gender=result.gender),
                        self.callback
                    )
                    result.seo_title = seo.title
                    result.seo_description = seo.description
                    result.seo_keywords = seo.keywords
            
            result.success = True
            self.callback(f"✅ {product_code} 처리 완료")
            
        except Exception as e:
            result.error_message = str(e)
            self.callback(f"❌ {product_code} 처리 실패: {e}")
        
        return result
    
    def _find_and_read_text(self, folder_path: Path) -> str:
        for txt_file in folder_path.glob("*.txt"):
            try:
                with open(txt_file, "r", encoding="utf-8") as f:
                    return f.read()
            except:
                try:
                    with open(txt_file, "r", encoding="cp949") as f:
                        return f.read()
                except:
                    pass
        return ""
    
    def _find_front_image(self, folder_path: Path) -> Optional[Path]:
        for ext in [".jpg", ".jpeg", ".png", ".webp"]:
            candidate = folder_path / f"5{ext}"
            if candidate.exists():
                return candidate
        img_ext = [".jpg", ".jpeg", ".png", ".webp"]
        images = sorted([f for f in folder_path.glob("*") if f.suffix.lower() in img_ext])
        return images[0] if images else None
    
    def _find_back_image(self, folder_path: Path) -> Optional[Path]:
        for ext in [".jpg", ".jpeg", ".png", ".webp"]:
            candidate = folder_path / f"6{ext}"
            if candidate.exists():
                return candidate
        return None
    
    def process_batch(self, work_folder: Path, product_folders: List[str], options: dict = None,
                      progress_callback: Callable[[int, int, str], None] = None) -> List[ProcessResult]:
        self.is_running = True
        self.should_stop = False
        results = []
        total = len(product_folders)
        
        for idx, folder_name in enumerate(product_folders):
            if self.should_stop:
                self.callback("⚠️ 사용자 요청으로 중단")
                break
            if progress_callback:
                progress_callback(idx + 1, total, folder_name)
            folder_path = work_folder / folder_name
            result = self.process_folder(folder_path, folder_name, options)
            results.append(result)
        
        self.is_running = False
        return results
    
    def stop(self):
        self.should_stop = True
